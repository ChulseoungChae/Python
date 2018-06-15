# -*- coding: utf-8 -*-

from __future__ import print_function
import sys
import time
import datetime

import requests
import json
import ketidatetime
import HTTP_POST_request
import openpyxl
from openpyxl import Workbook

import operator

import matplotlib.pyplot as plt
import numpy as np


""" 실제 작업에 필요한 query parameter 예 - 초기값 """
query_parameter = {
    "start" : "2014-06-01 00:00:00",
    "end": "2014-06-14 00:00:00",
    "aggregator" : "none",
    "metric" : "HanuriTN_00"
}

""" 쿼리 하려는 tag """
query_tags = {
    "content" : "gpsla"
}

def brush_args():
    usage = "\n usage : python %s {IP address} {port} " %sys.argv[0]
    usage += "{in_metric_name} "
    usage += "\n usage : python %s {IP address} {port} " %sys.argv[0]
    usage += "{in_metric_name} {start_day} {end_day}"
    print ( usage + '\n'  )

    _len = len(sys.argv)

    if _len < 5:
        print (" 추가 정보를 입력해 주세요. 위 usage 설명을 참고해 주십시오")
        print (" python 이후, 아규먼트 갯수 5개 필요 ")
        print (" 현재 아규먼트 %s 개가 입력되었습니다." %(_len) )
        print (" check *run.sh* file ")
        exit (" You need to input more arguments, please try again \n")

    _ip = sys.argv[1]
    _port = sys.argv[2]
    _in_met = sys.argv[3]
    _start = None
    _end = None

    if _len > 4:
        _start = sys.argv[4]
        _end = sys.argv[5]

    if ( len(_start) < 15 ):
        _start = ketidatetime._check_time_len(_start)
        _end = ketidatetime._check_time_len(_end)

    return _ip, _port, _in_met, _start, _end



def main_run(_date, dy=None, hrs=None, mins=None):

    if dy == 1: _t_scale = 24
    elif hrs == 1 : _t_scale = 1
    elif mins == 1 : _t_scale = 0.1
    if g_debug_off : print (_t_scale)

    query_parameter['start'] = _date
    query_parameter['end'] = ketidatetime.strday_delta(_date, _t_scale)
    query_parameter['aggregator'] = 'none'
    query_parameter['metric'] = in_metric

    _q_para = query_parameter
    _url = 'http://' + ip + ':' + port + '/api/query'
    if g_debug_on : print (_url, _q_para, query_tags)
    queryData = HTTP_POST_request.QueryData(_url, _q_para, query_tags)
    # 데이터 reading from DB, 스트링 데이터임

    _dictbuf, _dictlen = processingResponse(queryData)

    _graph_dict = WriteDataInExcel(_dictbuf, _dictlen)

    if _dictlen == 0 :
        print ("length is 0, please check it later, it is strange dps is empty")
        return None

    return _graph_dict

def processingResponse(in_data):
    '''openTSDB에서 전송받은, string 들을 dictionary로 변경하여
       dict 를 회신함. 형태를 변경하고, dict의 갯수를 알려줌'''

    _d = in_data
    _r = eval(_d.content)

    _l = len(_r)

    if g_debug_off :
        print ("... debugging print ... filename: %s" %__file__)
        print (type(_d.content), len(_d.content))
        print (type(_r), _l)

        if g_debug_off:
            print (_r[0]['metric'])
            print (_r[0]['tags'])
            print (_r[0]['dps'])

    return _r, _l
    # dict 와 length (딕셔너리 갯수) 리턴

def WriteDataInExcel(_dict_list, _len ):
    graphDict = {}
    # 빈 딕셔너리 생성
    _current_cnt = 0
    # 변수 _current_cnt을 0으로 선언

    global col_num

    count = 0
    for one_dict in _dict_list:
        _current_cnt += 1
        if g_debug_off: print ('\n')
        for k, v in one_dict.iteritems():
            row_num = 4
            if k == 'tags':
                _key = v['carid']
                if g_debug_off : print('carid : %s' %_key)
                sheet.cell(row=1, column=col_num, value='Carid')
                sheet.cell(row=1, column=col_num+1, value=_key)
                sheet.cell(row=2, column=col_num, value=str(v))
            if k == 'dps':
                count = len(v)
                if g_debug_off : print('num : %s' %count)
                if len(v) == 0 :
                    sheet.cell(row=row_num, column=col_num, value='No Data')
                for kt, pv in v.iteritems():
                    sheet.cell(row=row_num, column=col_num, value=kt)
                    sheet.cell(row=row_num, column=col_num+1, value=pv)
                    row_num += 1
        col_num += 3

        graphDict[_key] = count

    #print('col_num : %s \n' %col_num)
    #print(graphDict)

    return graphDict

def sum_processing(_idict, _odict):
    for k,v in _idict.iteritems():
        if g_debug_off: print("...marking line")
        if _odict.has_key(k): _odict[k] += _idict[k]
        else:  _odict[k] = _idict[k]

    if g_debug_off : print(_odict)
    return _odict


def sum_array(__sp):
    sortedArr = sorted(__sp.items(), key=operator.itemgetter(1, 0), reverse=True)
    print(sortedArr)
    global _second
    total_second = int(_second) + 2
    print('total second : %s\n' %total_second)

    n = 1
    for num in range(len(sortedArr)):
        if num == 0:
            print('rank %d [ carid : %s , value : %d ] %.2f%%' % (n, sortedArr[num][0], sortedArr[num][1], float(sortedArr[num][1]) / float(total_second)*100))
        else :
            if sortedArr[num][1] != sortedArr[num-1][1] :
                n += 1
            print('rank %d [ carid : %s , value : %d ] %.2f%%' %(n, sortedArr[num][0], sortedArr[num][1], float(sortedArr[num][1]) / float(total_second)*100))

    print('Total number : %s\n' %len(sortedArr))

    return sortedArr

def MakePyFile(MF):
    file = open("carid_list.py", 'w')

    carid = list()
    for k2, v2 in MF.iteritems():
        carid.append(k2)

    data1 = 'carid = %s\n' %carid

    file.write(data1)

    file.close()


def heatmap(__h):
    new_List = []
    for i in range(0, 20):
        new_List.append(list(np.zeros(30).astype(int)))

    #print(type(list_in_list))
    #print(type(new_List))
    #print('new_List : %s' % new_List)

    for j in range(1, 600):
        for k, v in __h.iteritems():
            if int(k) == j:
                x = 0
                y = 0
                x = j / 30
                y = j % 30
                #print(x, y)
                new_List[x-1][y-1] = v

    #print(new_List)

    hm = np.array(new_List)
    plt.axis()
    plt.title('title')
    plt.xlabel('X value')
    plt.ylabel('Y value')
    plt.imshow(hm, cmap='hot', interpolation='nearest', origin='lower')
    plt.colorbar()
    plt.show()


    return None

def plotgraph(__d):
    xList = list()
    yList = list()
    for k2, v2 in __d.iteritems():
        xList.append(k2)
        yList.append(v2)

    if g_debug_off: print('dict_numlen : %s\n' %len(__d))
    if g_debug_off: print('xList num : %s\n' %len(xList))

    part = len(__d) / 4
    if g_debug_off:print('part_len : %s' %part)

    #part1
    xList1 = xList[0:part]
    yList1 = yList[0:part]
    # part2
    xList2 = xList[part:part*2]
    yList2 = yList[part:part * 2]
    # part3
    xList3 = xList[part*2:part*3]
    yList3 = yList[part*2:part * 3]
    # part4
    xList4 = xList[part*3:]
    yList4 = yList[part*3:]
    if g_debug_off:print('xList1 : %s' % xList1)

    #first plot
    fig = plt.figure(figsize=(60, 40))
    plt.rc('font', size=7)
    xs = [i for i, _ in enumerate(xList1)]
    ax = fig.add_subplot(221)

    #print ([i for i, _ in enumerate(xList)],xList)
    #exit()

    plt.xticks([i for i, _ in enumerate(xList1)], xList1)
    rects = plt.bar(xs, yList1, align='center', width=0.3)
    for j, rect in enumerate(rects):
        ax.text(rect.get_x() + rect.get_width() / 2.0, 1.01 * rect.get_height(), str(yList1[j]), ha='center')
    plt.xlabel('car ID')
    plt.ylabel('counts')


    # second plot
    #fig2 = plt.figure(figsize=(60, 40))
    plt.rc('font', size=7)
    xs = [i for i, _ in enumerate(xList2)]
    ax= fig.add_subplot(222)
    plt.xticks([i for i, _ in enumerate(xList2)], xList2)
    rects = plt.bar(xs, yList2, align='center', width=0.3)
    for j, rect in enumerate(rects):
        ax.text(rect.get_x() + rect.get_width() / 2.0, 1.01 * rect.get_height(), str(yList2[j]), ha='center')
    plt.xlabel('car ID')
    plt.ylabel('counts')
    plt.rc('font', size=7)

    # third plot
    #fig3 = plt.figure(figsize=(60, 40))
    plt.rc('font', size=7)
    xs = [i for i, _ in enumerate(xList3)]
    ax = fig.add_subplot(223)
    plt.xticks([i for i, _ in enumerate(xList3)], xList3)
    rects = plt.bar(xs, yList3, align='center', width=0.3)
    for j, rect in enumerate(rects):
        ax.text(rect.get_x() + rect.get_width() / 2.0, 1.01 * rect.get_height(), str(yList3[j]), ha='center')
    plt.xlabel('car ID')
    plt.ylabel('counts')


    #Fourth.legend()
    #fig4 = plt.figure(figsize=(60, 40))
    plt.rc('font', size=7)
    xs = [i for i, _ in enumerate(xList4)]
    ax = fig.add_subplot(224)
    plt.xticks([i for i, _ in enumerate(xList4)], xList4)
    rects = plt.bar(xs, yList4, align='center', width=0.3)
    for j, rect in enumerate(rects):
        ax.text(rect.get_x() + rect.get_width() / 2.0, 1.01 * rect.get_height(), str(yList4[j]), ha='center')
    plt.xlabel('car ID')
    plt.ylabel('counts')


    plt.show()

#메인동작
if __name__ == "__main__":

    excel_file = Workbook()
    sheet = excel_file.active

    sum_dict = dict()

    g_debug_on = 1
    g_debug_off = 0

    start_time = time.time()
    start_datetime = datetime.datetime.now()

    ip, port, in_metric, q_start, q_end = brush_args()

    headers = {'content-type': 'application/json'}

    _date_on  = q_start
    _date_off = q_end

    _date_start = ketidatetime.str2datetime(q_start)
    _date_end = ketidatetime.str2datetime(q_end)

    _date_ =_date_end - _date_start
    _second = _date_.total_seconds()

    #print('date is : %s' %_date_)
    #print('date is : %d' %_second)

    col_num = 1

    while(_date_on != None):
        _r_dict = main_run(_date_on, mins=1)
        sum_dict = sum_processing(_r_dict, sum_dict)
        _date_on = ketidatetime.daydelta(_date_on, _date_off, mins=1)
    if g_debug_on: print ("=============================")
    if g_debug_on: print ("    sum dict result ")
    if g_debug_off: print (sum_dict)

    _sumarray = sum_array(sum_dict)

    MakePyFile(sum_dict)

    heatmap(sum_dict)

    plotgraph(sum_dict)

    #excel_file.save("carid.xlsx")

    run_time = time.time() - start_time
    print ( "\n\n [Main] Query time: %s ~ %s" % (q_start,q_end) )
    print ( " Run time: %.4f (sec)" % (run_time) )
