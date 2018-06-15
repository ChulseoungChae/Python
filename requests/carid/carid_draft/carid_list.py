# -*- coding: utf-8 -*-

from __future__ import print_function
import sys
import time
import datetime

import requests
import json
import ketidatetime
import HTTP_POST_request
import openpyxl
from openpyxl import Workbook

import matplotlib.pyplot as plt


""" 실제 작업에 필요한 query parameter 예 - 초기값 """
query_parameter = {
    "start" : "2014-06-01 00:00:00",
    "end": "2014-06-14 00:00:00",
    "aggregator" : "none",
    "metric" : "HanuriTN_00"
}

""" 쿼리 하려는 tag """
query_tags = {
    "content" : "gpsla"
}

def brush_args():
    usage = "\n usage : python %s {IP address} {port} " %sys.argv[0]
    usage += "{in_metric_name} "
    usage += "\n usage : python %s {IP address} {port} " %sys.argv[0]
    usage += "{in_metric_name} {start_day} {end_day}"
    print ( usage + '\n'  )

    _len = len(sys.argv)

    if _len < 5:
        print (" 추가 정보를 입력해 주세요. 위 usage 설명을 참고해 주십시오")
        print (" python 이후, 아규먼트 갯수 5개 필요 ")
        print (" 현재 아규먼트 %s 개가 입력되었습니다." %(_len) )
        print (" check *run.sh* file ")
        exit (" You need to input more arguments, please try again \n")

    _ip = sys.argv[1]
    _port = sys.argv[2]
    _in_met = sys.argv[3]
    _start = None
    _end = None

    if _len > 4:
        _start = sys.argv[4]
        _end = sys.argv[5]

    if ( len(_start) < 15 ):
        _start = ketidatetime._check_time_len(_start)
        _end = ketidatetime._check_time_len(_end)

    return _ip, _port, _in_met, _start, _end



def main_run(_date, dy=None, hrs=None, mins=None):

    if dy == 1: _t_scale = 24
    elif hrs == 1 : _t_scale = 1
    elif mins == 1 : _t_scale = 0.1
    if g_debug_off : print (_t_scale)

    query_parameter['start'] = _date
    query_parameter['end'] = ketidatetime.strday_delta(_date, _t_scale)
    query_parameter['aggregator'] = 'none'
    query_parameter['metric'] = in_metric

    _q_para = query_parameter
    _url = 'http://' + ip + ':' + port + '/api/query'
    if g_debug_on : print (_url, _q_para, query_tags)
    queryData = HTTP_POST_request.QueryData(_url, _q_para, query_tags)
    # 데이터 reading from DB, 스트링 데이터임

    _dictbuf, _dictlen = processingResponse(queryData)

    _graph_dict = WriteDataInExcel(_dictbuf, _dictlen)

    if _dictlen == 0 :
        print ("length is 0, please check it later, it is strange dps is empty")
        return None

    return _graph_dict

def processingResponse(in_data):
    '''openTSDB에서 전송받은, string 들을 dictionary로 변경하여
       dict 를 회신함. 형태를 변경하고, dict의 갯수를 알려줌'''

    _d = in_data
    _r = eval(_d.content)

    _l = len(_r)

    if g_debug_off :
        print ("... debugging print ... filename: %s" %__file__)
        print (type(_d.content), len(_d.content))
        print (type(_r), _l)

        if g_debug_off:
            print (_r[0]['metric'])
            print (_r[0]['tags'])
            print (_r[0]['dps'])

    return _r, _l
    # dict 와 length (딕셔너리 갯수) 리턴

def WriteDataInExcel(_dict_list, _len ):
    graphDict = {}
    # 빈 딕셔너리 생성
    dict_array = []
    # 빈 리스트 생성
    _current_cnt = 0
    # 변수 _current_cnt을 0으로 선언

    all_rows = sheet.rows
    col_num = 1

    count = 0
    for one_dict in _dict_list:
        _current_cnt += 1
        if g_debug_off: print ('\n')
        for k, v in one_dict.iteritems():
            row_num = 4
            sheet.cell(row=1, column=col_num, value=str(k))
            sheet.cell(row=2, column=col_num, value=str(v))
            if k == 'tags': _key = v['carid']
            elif k == 'dps':
                count = len(v)
                if g_debug_off : print('num : %s' %count)
            col_num += 2

        graphDict[_key] = count

    return graphDict

def sum_processing(_idict, _odict):
    for k,v in _idict.iteritems():
        if g_debug_off: print("...marking line")
        if _odict.has_key(k): _odict[k] += _idict[k]
        else:  _odict[k] = _idict[k]
    return _odict

def plotgraph(__d):
    plt.plot([0, 1,2, 7], [1, 1,2, 7], label='count')
    plt.plot([0,2,5], [1,3,5], marker='o', linestyle='--', color='b', label='tcount')
    plt.xlabel('car ID')
    plt.ylabel('counts')
    plt.legend()
    plt.show(block=False)
    plt.show()

#메인동작
if __name__ == "__main__":

    excel_file = Workbook()
    sheet = excel_file.active

    sum_dict = {}

    g_debug_on = 1
    g_debug_off = 0

    start_time = time.time()
    start_datetime = datetime.datetime.now()

    ip, port, in_metric, q_start, q_end = brush_args()

    headers = {'content-type': 'application/json'}

    _date_on  = q_start
    _date_off = q_end

    while(_date_on != None):
        _r_dict = main_run(_date_on, mins=1)
        sum_dict = sum_processing(_r_dict, sum_dict)
        _date_on = ketidatetime.daydelta(_date_on, _date_off, mins=1)
    if g_debug_on: print ("=============================")
    if g_debug_on: print ("    sum dict result ")
    if g_debug_off: print (sum_dict)

    plotgraph(sum_dict)

    excel_file.save("carid.xlsx")

    run_time = time.time() - start_time
    print ( "\n\n [Main] Query time: %s ~ %s" % (q_start,q_end) )
    print ( " Run time: %.4f (sec)" % (run_time) )
