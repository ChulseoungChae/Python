# -*- coding: utf-8 -*-

from __future__ import print_function
import sys
import os
import csv
import datetime
import matplotlib.pyplot as plt
import glob
from collections import Counter
import operator
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Color


def str2datetime(dt_str):
    return datetime.datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")


def graph(k_data, v_data):
    _month = ['201406', '201407', '201408', '201409', '201410', '201411', '201412',]

    hm_x = np.array(k_data)
    hm_y = np.array(_month)
    hm_z = np.array(v_data)
    plt.figure(figsize=(12, 8))
    plt.axis()
    # plt.title('title')
    plt.xlabel('Latitude')
    plt.ylabel('Longitude')
    plt.xticks(hm_x)
    plt.xticks(fontsize=7)  # 디폴트값:10
    plt.yticks(hm_y)
    plt.yticks(fontsize=7)  # 디폴트값:10
    plt.pcolor(hm_x, hm_y, hm_z, cmap='cool')
    # plt.imshow(hm_z, extent=[hm_x.min(), hm_x.max(), hm_y.min(), hm_y.max()], cmap='OrRd', interpolation='nearest', origin='lower')
    for y in range(hm_z.shape[0] - 1):
        for x in range(hm_z.shape[1] - 1):
            plt.text(lon_list[x] + (lat / 2), lat_list[y] + (lat / 2), np_num[y, x], horizontalalignment='center',
                     verticalalignment='center')
    plt.colorbar()
    fig = plt.gcf()
    plt.show()
    fig.savefig(data_list[2][t] + 'heatmap.png')

'''
def User_input(month_dict):
    user_input = input("데이터를 확인할 월을 입력하세요(ex. 201406)(모든월: all 입력 ):")
    if user_input in month_dict:
        data = month_dict[user_input]
    else:
        print('다시 입력해주세요')
        User_input()
    return data
'''

def dict2list(_data):
    key_list = []
    value_list = []

    sorted_data = sorted(_data.items(), key=operator.itemgetter(0))
    #print(sorted_data)
    #print(len(sorted_data))
    for i in range(len(sorted_data)):
        key_list.append(sorted_data[i][0])
        value_list.append(sorted_data[i][1])

    #print(key_list)
    #print(value_list)
    #print(len(key_list))
    #print(len(value_list))

    return key_list, value_list


def Excel(dict):
    print('make excel file...')
    excel_file = Workbook()
    sheet = excel_file.active

    all_key, all_value = dict2list(dict['all'])
    for i in range(len(all_key)):
        sheet.cell(row=1, column=2, value='all').fill = PatternFill(patternType='solid', fgColor=Color('4785F0'))
        sheet.cell(row=i + 2, column=1, value=all_key[i]).fill = PatternFill(patternType='solid', fgColor=Color('4785F0'))
        sheet.cell(row=i + 2, column=2, value=all_value[i]).fill = PatternFill(patternType='solid', fgColor=Color('FFFF00'))

    _count = 1
    for key in dict:
        k_data, v_data = dict2list(dict[key])
        _count +=1
        if key != 'all':
            for j in range(len(k_data)):
                for k in all_key:
                    if k_data[j] == k:
                        sheet.cell(row=1, column=_count, value=key).fill = PatternFill(patternType='solid', fgColor=Color('4785F0'))
                        sheet.cell(row=all_key.index(k)+2, column=_count, value=v_data[j]).fill = PatternFill(patternType='solid', fgColor=Color('FFFF00'))

    excel_file.save('total_count.xlsx')
    print('Finish')


if __name__ == "__main__":
    dir = '/home/data/data_1/Hanuritien/hanuritien/'
    month_list = glob.glob(dir+'*')
    month_list.sort()

    month_dict = {'201406':0, '201407':0, '201408':0, '201409':0, '201410':0, '201411':0, '201412':0, 'all':0}
    for month in month_list:
        carid_list = glob.glob(month+'/*')
        carid_list.sort()
        _dict = {}
        for carid in carid_list:
            day_list = glob.glob(carid+'/*')
            day_list.sort()
            _carid = carid[-3:]
            _num = 0
            for day_file in day_list:
                spd_list = []
                if _carid == '001':
                    with open(day_file, mode='r') as f:
                        reader = csv.reader(f)
                        for row in reader:
                            spd_list.append(row[5])
                        spd_list[0:1] = []
                        _num += len(spd_list)

            print('%s is ok' %carid)
            _dict[_carid] = _num

        if month[-6:] in month_dict:
            month_dict[month[-6:]] = _dict
        #print(month_dict)

    all_month = Counter(month_dict['201406'])+Counter(month_dict['201407'])+Counter(month_dict['201408'])+Counter(month_dict['201409'])+Counter(month_dict['201410'])+Counter(month_dict['201411'])+Counter(month_dict['201412'])
    month_dict['all'] = all_month

    #_data = User_input(month_dict)
    #k_data, v_data = dict2list(_data)

    Excel(month_dict)

    # graph(k_data, v_data)
