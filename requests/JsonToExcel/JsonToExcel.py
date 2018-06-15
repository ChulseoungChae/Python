# -*- coding: utf-8 -*-
from __future__ import print_function #파이썬2에서 파이썬3 프린트문을 다루기위해
import sys #사용자가 입력한 인자값을 다루기 위해
import json #json파일을 다루기 위해
import openpyxl #엑셀파일을 사용하기 위해
from openpyxl import Workbook #엑셀파일 워크시트 사용을 위해
from openpyxl.styles import PatternFill, Color #셀에 색상을 입히기 위해 사용

#import pandas

# 사용자가 입력한 json파일 이름을 리턴하는 함수
def input_filename():
    usage = "\n usage : python %s {FileName} " %sys.argv[0]
    print ( usage + '\n'  )
    # usage를 출력

    _len = len(sys.argv)
    # 파이썬 파일 이름 뒤에 인자의 길이를 저장

    if _len < 1:
        print (" 파일 이름을 입력해주세요 ")
        print (" check *run.sh* file ")
        exit (" You need to input File name, please try again \n")
    # 파이썬 파일 이름 뒤에 입력한 인자가 없으면 프린트문을 출력하고 종료

    file_name = sys.argv[1]
    # 파이썬 파일이름 뒤에 입력한 json파일 이름을 저장

    return file_name
    # json파일 이름을 리턴
    

# 엑셀에 json파일 내용을 입력하는 함수
def WriteDataInExcel(_dict):
    col_num = 1
    # column값 선언
    
    for one_dict in _dict:
        for k, v in one_dict.iteritems():
            if k == 'period':
                sheet.cell(row=1, column=1, value=k).fill = PatternFill(patternType='solid', fgColor=Color('4785F0'))
                sheet.cell(row=1, column=2, value=v).fill = PatternFill(patternType='solid', fgColor=Color('FFFF00'))
            # 딕셔너리 key가 period이면 1,1셀에 period입력, 1,2에 period value값 입력하고 셀에 색 입힌다

            if k == 'data':
            # 딕셔너리 key가 data이면     
                for _list in v:
                    for kt, pv in _list.iteritems():
                        if kt == 'carid':
                            sheet.cell(row=3, column=col_num, value=kt).fill = PatternFill(patternType='solid', fgColor=Color('FF0000'))
                            sheet.cell(row=3, column=col_num+1, value=pv).fill = PatternFill(patternType='solid', fgColor=Color('FFFF00'))
                        # carid일때 셀에 carid와 그 value값 입력
                            
                        if kt == 'count':
                            sheet.cell(row=4, column=col_num, value=kt).fill = PatternFill(patternType='solid', fgColor=Color('FF0000'))
                            sheet.cell(row=4, column=col_num+1, value=pv).fill = PatternFill(patternType='solid', fgColor=Color('FFFF00'))
                        # count일때 셀에 count와 그 value값 입력
                            
                        if kt == 'time':
                            row_num = 6
                            sheet.cell(row=row_num, column=col_num, value=kt).fill = PatternFill(patternType='solid', fgColor=Color('FF0000'))
                            for pv_list in pv:
                                sheet.cell(row=row_num, column=col_num+1, value=pv_list).fill = PatternFill(patternType='solid', fgColor=Color('FFFF00'))
                                row_num += 1
                        # time일때 셀에 time과 그 value값 입력

                    col_num += 4
                    #column 데이터가 입력되는 행부터 4칸 띄우기

        #print(col_num)
        
        col_num +=3
        # 행 3칸 띄우기

        for k, v in one_dict.iteritems():
            if k == "timestamp_count":
                sheet.cell(row=4, column=col_num, value=k).fill = PatternFill(patternType='solid', fgColor=Color('FF0000'))
                row_num = 6
                for td_k, td_v in v.iteritems():
                    sheet.cell(row=row_num, column=col_num, value=td_k).fill = PatternFill(patternType='solid', fgColor=Color('00FF00'))
                    sheet.cell(row=row_num, column=col_num+1, value=td_v).fill = PatternFill(patternType='solid', fgColor=Color('CCEEFF'))
                    row_num += 1
                    if row_num == 15006:
                        row_num = 6
                        col_num += 4
        # timestamp_count는 1행에 15000개씩 나눠서 데이터를 입력
        # 엑셀 2007이후 쓸수 있는 최대 행열 : 1048576행(약 백만 개) x 16384열(약 만육천 개)

        #print(col_num)

    return None


if __name__ == "__main__":
    fn = input_filename() # json파일 이름을 저장
    
    excel_file = Workbook() # 빈 엑셀파일 생성
    sheet = excel_file.active # 현재 로드된 엑셀 sheet 지정

    json_data=open(fn) # json파일 오픈
    dict = json.load(json_data) # json파일을 디코드해서 리스트형식으로 바꿈
    #print(dict[0]['data'][0]['carid'])

    WriteDataInExcel(dict)  # 함수호출하여 엑셀에 데이터 입력

    excel_file.save(fn + '.xlsx') # 작업한 엑셀파일 저장

    #pandas.read_json(fn).to_excel(fn + '.xlsx')
